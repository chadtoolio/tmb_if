# payroll input portal office access only
# TODO a standard practice that can be done by any manager.
# TODO input to the gui will prompt for next employee alphabetically
# TODO append data in spreadsheet via GUI interface
# TODO password true false statements added for next step
# TODO calculate taxes eliminating the need for adp
# TODO overides and notifications for irregular items require manager approval
# TODO pay amount cannot be seen by managers
# TODO banking wire transfers for payroll to employees
# TODO input birthdays for christmas bonus calculationfrom openpyxl.workbook import Workbook
from openpyxl import load_workbook


thrwb = load_workbook('CTMTHRS.xlsx')
bswb = load_workbook('ctmcfimp.xlsx')
mproll = load_workbook('Clone_payroll.xlsx')
trey_sh = load_workbook('trey-boensch.xlsx')
ryan_sh = load_workbook('ryan-bowers.xlsx')
austin_sh = load_workbook('austin-calvert.xlsx')
anthony_sh = load_workbook('anthony-cucaz.xlsx')
papa_sh = load_workbook('roosevelt-finley.xlsx')
rob_sh = load_workbook('rob-follmar.xlsx')
chip_sh = load_workbook('chip-myers.xlsx')
Proll_GUI_wb = load_workbook('TMB-Payroll-GUI.xlsx')
