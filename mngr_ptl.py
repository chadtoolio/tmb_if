# TODO add employee gui
# TODO add code to GUI "if employee is filling a tech position, create list of sliding pay scales with if statemnts



import openpyxl


from openpyxl import  Workbook, load_workbook
# define what workbook or file you want to load
wb = load_workbook("test_db.xlsx")
ws = wb.active

# print the value of an individual cell
print(ws['B4'].value)

# change the contents of any cell, could be a formula as well
ws['B25'].value = "=sum(b13+c13)"
wb.save("test_db.xlsx")

# get sheetnames of loaded workbook
print(wb.sheetnames)

# creat new sheet
wb.create_sheet("john mechanic")

# get sheetnames of loaded workbook
print(wb.sheetnames)

